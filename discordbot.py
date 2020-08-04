#!/root/.pyenv/versions/3.8.2/bin/ python3
# インストールした discord.py を読み込む
import discord
from discord.ext import tasks
#from datetime import datetime
import re
from datetime import datetime, timedelta, timezone
from time import sleep
import gspread
#ServiceAccountCredentials：Googleの各サービスへアクセスできるservice変数を生成します。
from oauth2client.service_account import ServiceAccountCredentials 
import json
import configparser
import time
import math
from discord.ext import commands
import pickle
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook
import jaconv
import asyncio

# 自分のBotのアクセストークンを指定
TOKEN = 'NjQ1NDcyNDg0NDQzMDI5NTMw.XxlVfw.BKamFS6DauMbhnrGeB9x5UIXNv0'
lot_server_id = 561824358121734161 # サーバIDを指定(トロイメライ用)
ID_CHANNEL_REACT = 695812943236956190  # 簡易入力用チャンネルID（事前設定用）
lot_channel_id = 561837260710871050 # 3凸報告用のチャンネルを指定
ID_EMOJI_RESERV = [672857988423352411, 672858267806072832, 672858291554222176, 672858310952878090]
ID_EMOJI_NEXT_RESERV = [675498654211244032, 675498675774029826, 675498696338702338, 675498713711640604]
yoyaku_id = 692003669712961586
ID_CHANNEL_RESERV = 695813057808564244 # スタンプ出力用チャンネルID（事前設定用）※ちょいちょい修正入る
ID_CHANNEL_MAIN = 695812943236956190 # 入力用チャンネルID（事前設定用）


# 自分のBotのアクセストークンを指定◆テストサーバ◆
#TOKEN = 'NjY4NDY2MzU1OTk3MjQ1NDQx.XlEcyg.EQ4A2N8I--SffrN7MPnsMVmA6v0'
#lot_server_id = 630404930381021214  # サーバIDを指定◆テストサーバ◆
#ID_CHANNEL_REACT = 668464753621860352  # 簡易入力用チャンネルID（事前設定用）◆テストサーバ◆
#ID_EMOJI_RESERV = [668473808549707799, 668473869316784148, 668473887897550876, 668473908420149248] #◆テストサーバ◆
#ID_EMOJI_NEXT_RESERV = [680751766123446283, 680751781466341396, 680751798973104220, 680751812391075848] #◆テストサーバ◆
#yoyaku_id = 668464625917886484 #◆テストサーバ◆
#ID_CHANNEL_RESERV = 668464753621860352 # 簡易入力用チャンネルID（事前設定用）#◆テストサーバ◆
#lot_channel_id = 653226713764855808 # 3凸報告用のチャンネルを指定（テスト用）
#ID_CHANNEL_MAIN = 668464247751311361 # 入力用チャンネルID（事前設定用）#◆テストサーバ◆

#-------------------------------以降鯖毎の変更点なし---------------------------------------------

#★★★スプシ用変数★★★
#2つのAPIを記述しないとリフレッシュトークンを3600秒毎に発行し続けなければならない
#scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
#認証情報設定
#ダウンロードしたjsonファイル名をクレデンシャル変数に設定（秘密鍵、Pythonファイルから読み込みしやすい位置に置く）
#credentials = ServiceAccountCredentials.from_json_keyfile_name('graphite-shell-264116-3c60f1b59991.json', scope)
#OAuth2の資格情報を使用してGoogle APIにログインします。
#gc = gspread.authorize(credentials)
#共有設定したスプレッドシートキーを変数[SPREADSHEET_KEY]に格納する。
#SPREADSHEET_KEY = '1g36QIp8KZFGNwBkJG994XE7y4lOv9ojXUGmwmAaRGUM'
#共有設定したスプレッドシートのシート1を開く
#worksheet = gc.open_by_key(SPREADSHEET_KEY).sheet1

# 接続に必要なオブジェクトを生成
client = discord.Client()
JST = timezone(timedelta(hours=+9), 'JST')
DELAY_S = 3  # メッセージ削除までの時間（短）
DELAY_M = 30  # メッセージ削除までの時間（中）
DELAY_L = 60  # メッセージ削除までの時間（長）
bossindex = 0
bosshp = [700,900,1300,1500,2000]

#★追加変数★
Emoji_Command = ["物理凸", "物理持越し", "魔法凸", "魔法持越し"]
Emoji_Next_Command = ["次週物理凸", "次週物理持越し", "次週魔法凸", "次週魔法持越し"]
Message_Boss_Reaction = []  # ボス凸リアクション用メッセージオブジェクト配列
boss1_mage = []
boss1_physic = []
boss1_carry = []
boss2_mage = []
boss2_physic = []
boss2_carry = []
boss3_mage = []
boss3_physic = []
boss3_carry = []
boss4_mage = []
boss4_physic = []
boss4_carry = []
boss4_carry = []
boss5_mage = []
boss5_physic = []
boss5_carry = []
boss1_total = []
boss2_total = []
boss3_total = []
boss4_total = []
boss5_total = []
next_boss1_mage = []
next_boss1_physic = []
next_boss1_carry = []
next_boss2_mage = []
next_boss2_physic = []
next_boss2_carry = []
next_boss3_mage = []
next_boss3_physic = []
next_boss3_carry = []
next_boss4_mage = []
next_boss4_physic = []
next_boss4_carry = []
next_boss5_mage = []
next_boss5_physic = []
next_boss5_carry = []
next_boss1_total = []
next_boss2_total = []
next_boss3_total = []
next_boss4_total = []
next_boss5_total = []
#damage_channel_id = 672858958926446592 #同時凸報告チャンネル
#input_channel_id = 695812943236956190 #入力反映用チャンネル
#emoji_owari = 652105400333762566 #確定用絵文字
genHP = 0
#★ボス周回数
round = 1
Boss_Confirm_Flag = False
Boss_Damage = 0
OutputCount = 0
playData = []
ClabatoMode = False
mayopote = 730072376074698873
dennys = 730086259824918761

# 持ち越しダメージ計算
async def rollover_simulate(message, msg_content):
    current_hp = 0
    expect_dmg = 0
    is_current_hp = False  # 現在のHP判定
    is_expected_dmg = False  # 想定ダメージ判定
    for i in msg_content:
        # ボス番号読み取って該当ボスのフラグを立てる
        if not is_expected_dmg and re.match('\\d', i):
            is_current_hp = True
            current_hp = current_hp * 10 + int(i)
        if is_expected_dmg and re.match('\\d', i):
            expect_dmg = expect_dmg * 10 + int(i)
        # 現HP読取後の区切り以降は想定ダメージとして読取
        if is_current_hp and not is_expected_dmg and (i == '-' or i == 'ー' or i == '－' or i == ' ' or i == '　'):
            is_expected_dmg = True
    print(current_hp)
    if not expect_dmg:  # 想定ダメージが入力されていない場合、一定秒持越に必要なダメージを出す
        i = 70
        reply = '```持越秒数に対して必要なダメージ一覧\n\n'
        while True:
            suggested_dmg = current_hp / (1 - i / 90)
            reply += f"{i + 20}s：{int(suggested_dmg)} 万ダメージ\n"
            i -= 5
            if i < 0:
                break
        reply += "```"
        await reply_and_delete(message, reply, DELAY_L)
        return
    elif expect_dmg < current_hp:
        reply = "そのダメージだと倒しきれません"
        await reply_and_delete(message, reply, DELAY_S)
        return

    rolled_time = 90 * (1 - current_hp / expect_dmg) + 20
    if rolled_time >= 90:
        rolled_time = 90

    reply = "予想される持越時間は " + str(rolled_time) + "秒です"
    await reply_and_delete(message, reply, DELAY_M)
    return

# 一定時間でログクリア
async def reply_and_delete(message, txt, delay_sec):
    # 簡易入力利用時は、入力元にもリプライ
    if message.author.id == client.user.id and message.mentions:
        global ID_CHANNEL_REACT
        channel = client.get_channel(ID_CHANNEL_REACT)
        tmp_msg = await channel.send(txt)
        #await tmp_msg.delete(delay=delay_sec)

    tmp_msg = await message.channel.send(txt)
    #await tmp_msg.delete(delay=delay_sec)

# ★凸予約用項目を展開★
async def init_reserv_channel():
    global ID_CHANNEL_MAIN
    global ID_CHANNEL_RESERV
    global Message_Boss_Reaction
    global ID_EMOJI_RESERV
    global ID_EMOJI_NEXT_RESERV
    global Emoji_Command
    global Emoji_Next_Command

    reserv_channel = client.get_channel(ID_CHANNEL_RESERV)

    if ID_CHANNEL_RESERV and ID_CHANNEL_MAIN:  # 初期設定済の場合、コマンドチャンネルに初期化メッセージ
        orig_channel = client.get_channel(ID_CHANNEL_MAIN)
    elif ID_CHANNEL_RESERV:  # 簡易入力チャンネルのみ設定されている場合、そこに初期化メッセージ
        orig_channel = client.get_channel(ID_CHANNEL_RESERV)
    else:  # 初期設定されておらず、コマンドからの設定でもない場合は終了
        return

    for emoji_id in ID_EMOJI_RESERV :
        if emoji_id is None:
            return

    for emoji_next_id in ID_EMOJI_NEXT_RESERV :
        if emoji_next_id is None:
            return

    # リアクション場所設置には時間かかるので、作業開始を伝える
    init_msg = await orig_channel.send("凸予約用チャンネルの初期化中です、しばらくお待ち下さい")

    # 再実行用にボス凸リアクション用メッセージオブジェクト配列の初期化
    Message_Boss_Reaction = []

    # 指定IDのチャンネルに簡易入力用項目を展開
    reply = "----------------------------------------\n"
    for i in range(4):
        reply += Emoji_Command[i] + "入力：" + str(client.get_emoji(ID_EMOJI_RESERV[i])) + "\n"
    reply += "----------------------------------------"
    await reserv_channel.send(reply)
    # 5ボス分 メッセージを投稿（0から始まるので＋１してボス名投稿）
    for i in range(5):
        tmp_msg_reserv = await reserv_channel.send(str(i + 1) + "ボス物理　｜" + str(i + 1) + "ボス魔法")
        for emoji_ID in ID_EMOJI_RESERV :
            await tmp_msg_reserv.add_reaction(client.get_emoji(emoji_ID))
        Message_Boss_Reaction.append(tmp_msg_reserv)
    reply = "----------------------------------------\n"
    await reserv_channel.send(reply)
    #次週凸予約
    for i in range(4):
        reply += Emoji_Next_Command[i] + "入力：" + str(client.get_emoji(ID_EMOJI_NEXT_RESERV[i])) + "\n"
    reply += "----------------------------------------"
    await reserv_channel.send(reply)
    # 5ボス分 メッセージを投稿（0から始まるので＋１してボス名投稿）
    for i in range(5):
        tmp_msg_reserv = await reserv_channel.send(str(i + 1) + "ボス物理次｜" + str(i + 1) + "ボス魔法次")
        for emoji_ID in ID_EMOJI_NEXT_RESERV :
            await tmp_msg_reserv.add_reaction(client.get_emoji(emoji_ID))
        Message_Boss_Reaction.append(tmp_msg_reserv)
    reply = "----------------------------------------"
    await reserv_channel.send(reply)    
    # リアクション対象メッセージをつくりおわったので、告知メッセージは閉じる
    await init_msg.delete()

# ★代理コマンド発行★
async def send_command_by_reaction(reply, reaction, orig_user):
    # orig_userへのメンションをつけてコマンドを代理投稿する
    reply += ' ' + orig_user.mention

    # コマンド用チャンネルが設定されていなかったら代理コマンドはリアクションを受け付けたチャンネルに投稿する
    if ID_CHANNEL_MAIN is not None:
        channel = client.get_channel(ID_CHANNEL_MAIN)  # 代理コマンドをコマンド用チャンネルに行う
        # リアクション投稿を受け付けた元チャンネルをメンションする
        #reply += ' ' + reaction.message.channel.mention
    else:
        channel = reaction.message.channel  # 未設定のためリアクションのあったメッセージのあるチャンネルにする

    # 代理投稿
    await channel.send(reply)

#★プレイヤークラス★
class PlayerData:
    def __init__(self, user, Boss_Damage, Boss_Confirm_Flag):
        self.user = user  # Discord.user オブジェクト
        self.Boss_Damage = Boss_Damage #ボスのダメージ
        self.Boss_Confirm_Flag = Boss_Confirm_Flag #ダメ確定済みフラグ
    # Discord.user Object
    def user(self):
        return self.user
    # ボスダメージ取得
    def add_Boss_Damage(self,int_boss_damage):
        self.Boss_Damage = int_boss_damage
        return    
    # ボス確定フラグ立てる
    def add_Boss_Confirm_Flag(self):
        self.Boss_Confirm_Flag = True
        return

# ★リアクション処理★
@client.event
async def on_raw_reaction_add(payload):
    global boss1_mage
    global boss1_physic
    global boss1_carry
    global boss2_mage
    global boss2_physic
    global boss2_carry
    global boss3_mage
    global boss3_physic
    global boss3_carry
    global boss4_mage
    global boss4_physic
    global boss4_carry
    global boss5_mage
    global boss5_physic
    global boss5_carry
    global boss1_total
    global boss2_total
    global boss3_total
    global boss4_total
    global boss5_total
    global next_boss1_mage
    global next_boss1_physic
    global next_boss1_carry
    global next_boss2_mage
    global next_boss2_physic
    global next_boss2_carry
    global next_boss3_mage
    global next_boss3_physic
    global next_boss3_carry
    global next_boss4_mage
    global next_boss4_physic
    global next_boss4_carry
    global next_boss5_mage
    global next_boss5_physic
    global next_boss5_carry
    global next_boss1_total
    global next_boss2_total
    global next_boss3_total
    global next_boss4_total
    global next_boss5_total
    global ClabatoMode
    global boss_index
    global ID_CHANNEL_RESERV

    channel = client.get_channel(payload.channel_id)  # チャンネル取得
    msg = await channel.fetch_message(payload.message_id)  # リアクションの付いたメッセージ取得
    user = client.get_user(payload.user_id)  # リアクションをつけたユーザーを取得

    #スタンプのメッセージIDを取得する
    if payload.message_id == 705013433031983166:
        boss_index = 0
    elif payload.message_id == 705013441668055351:
        boss_index = 1
    elif payload.message_id == 705013450509647944:
        boss_index = 2
    elif payload.message_id == 705013459183730718:
        boss_index = 3
    elif payload.message_id == 705013467320549396:
        boss_index = 4
    elif payload.message_id == 705013478272008244:
        boss_index = 5
    elif payload.message_id == 705013486543044710:
        boss_index = 6
    elif payload.message_id == 705013495267328030:
        boss_index = 7
    elif payload.message_id == 705013503731302442:
        boss_index = 8
    elif payload.message_id == 705013512916828192:
        boss_index = 9

    #★★★スタンプのメッセージIDを取得する(テスト用)★★★
#    if payload.message_id == 706481010119933954:
#        boss_index = 0
#    elif payload.message_id == 706481018969784340:
#        boss_index = 1
#    elif payload.message_id == 706481026993618964:
#        boss_index = 2
#    elif payload.message_id == 706481035805982730:
#        boss_index = 3
#    elif payload.message_id == 706481044005716058:
#        boss_index = 4
#    elif payload.message_id == 706481056915914752:
#        boss_index = 5
#    elif payload.message_id == 706481065778348076:
#        boss_index = 6
#    elif payload.message_id == 706481073944526919:
#        boss_index = 7
#    elif payload.message_id == 706481082421477406:
#        boss_index = 8
#    elif payload.message_id == 706481091116007424:
#        boss_index = 9

    # ボス凸対応リアクションを確認して、対象であったら対応コマンドを発行
    for react_index, emoji_id in enumerate(ID_EMOJI_RESERV):
        if payload.emoji == client.get_emoji(emoji_id):
            is_mage = react_index // 2
            is_rolled = react_index % 2
            reply = str(boss_index + 1)
            if is_mage:
                reply += "魔予約"
            else:
                reply += "物予約"
            if is_rolled:
                reply += "持越し"
            await send_command_by_reaction(reply, payload.emoji, user)  # 代理コマンド
            if boss_index == 0:
                if react_index == 0:
                    boss1_physic.append(user.name)
                    boss1_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    boss1_carry.append(user.name)
                    boss1_total.append(user.name)
                elif react_index == 2:
                    boss1_mage.append(user.name)
                    boss1_total.append(user.name)
            elif boss_index == 1:
                if react_index == 0:
                    boss2_physic.append(user.name)
                    boss2_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    boss2_carry.append(user.name)
                    boss2_total.append(user.name)
                elif react_index == 2:
                    boss2_mage.append(user.name)
                    boss2_total.append(user.name)
            elif boss_index == 2:
                if react_index == 0:
                    boss3_physic.append(user.name)
                    boss3_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    boss3_carry.append(user.name)
                    boss3_total.append(user.name)
                elif react_index == 2:
                    boss3_mage.append(user.name)
                    boss3_total.append(user.name)
            elif boss_index == 3:
                if react_index == 0:
                    boss4_physic.append(user.name)
                    boss4_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    boss4_carry.append(user.name)
                    boss4_total.append(user.name)
                elif react_index == 2:
                    boss4_mage.append(user.name)
                    boss4_total.append(user.name)
            elif boss_index == 4:
                if react_index == 0:
                    boss5_physic.append(user.name)
                    boss5_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    boss5_carry.append(user.name)
                    boss5_total.append(user.name)
                elif react_index == 2:
                    boss5_mage.append(user.name)
                    boss5_total.append(user.name)
    for react_index, emoji_id in enumerate(ID_EMOJI_NEXT_RESERV):
        if payload.emoji == client.get_emoji(emoji_id):
            is_mage = react_index // 2
            is_rolled = react_index % 2
            reply = str(boss_index - 4)
            if is_mage:
                reply += "魔予約(次)"
            else:
                reply += "物予約(次)"
            if is_rolled:
                reply += "持越し(次)"
            await send_command_by_reaction(reply, payload.emoji, user)  # 代理コマンド
            if boss_index == 5:
                if react_index == 0:
                    next_boss1_physic.append(user.name)
                    next_boss1_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    next_boss1_carry.append(user.name)
                    next_boss1_total.append(user.name)
                elif react_index == 2:
                    next_boss1_mage.append(user.name)
                    next_boss1_total.append(user.name)
            elif boss_index == 6:
                if react_index == 0:
                    next_boss2_physic.append(user.name)
                    next_boss2_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    next_boss2_carry.append(user.name)
                    next_boss2_total.append(user.name)
                elif react_index == 2:
                    next_boss2_mage.append(user.name)
                    next_boss2_total.append(user.name)
            elif boss_index == 7:
                if react_index == 0:
                    next_boss3_physic.append(user.name)
                    next_boss3_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    next_boss3_carry.append(user.name)
                    next_boss3_total.append(user.name)
                elif react_index == 2:
                    next_boss3_mage.append(user.name)
                    next_boss3_total.append(user.name)
            elif boss_index == 8:
                if react_index == 0:
                    next_boss4_physic.append(user.name)
                    next_boss4_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    next_boss4_carry.append(user.name)
                    next_boss4_total.append(user.name)
                elif react_index == 2:
                    next_boss4_mage.append(user.name)
                    next_boss4_total.append(user.name)
            elif boss_index == 9:
                if react_index == 0:
                    next_boss5_physic.append(user.name)
                    next_boss5_total.append(user.name)
                elif react_index == 1 or react_index == 3:
                    next_boss5_carry.append(user.name)
                    next_boss5_total.append(user.name)
                elif react_index == 2:
                    next_boss5_mage.append(user.name)
                    next_boss5_total.append(user.name)
    # 処理をしたらリアクションは消してやる（同コマンド代理を何度もできるように）
    if payload.channel_id == ID_CHANNEL_RESERV:
        await msg.remove_reaction(payload.emoji, user)

    #★ダメージ確定フラグを立てる★
#    if reaction.emoji == client.get_emoji(emoji_owari):
#        if reaction.message.channel.id == lot_channel_id:
#            for p in playData:
#                if p.user == reaction.message.author.name:
#                    p.add_Boss_Confirm_Flag()

# 5am定時処理
async def rollover_by5am():
    global Message_Log_Main
    global Flg_Sleep
    global Flg_Demo
    global OutputCount
    global yoyaku_id
    global round
    global boss1_mage
    global boss1_physic
    global boss2_mage
    global boss2_physic
    global boss3_mage
    global boss3_physic
    global boss4_mage
    global boss4_physic
    global boss5_mage
    global boss5_physic
    global boss1_total
    global boss2_total
    global boss3_total
    global boss4_total
    global boss5_total
    global next_boss1_mage
    global next_boss1_physic
    global next_boss2_mage
    global next_boss2_physic
    global next_boss3_mage
    global next_boss3_physic
    global next_boss4_mage
    global next_boss4_physic
    global next_boss5_mage
    global next_boss5_physic
    global boss1_carry
    global next_boss1_carry
    global boss2_carry
    global next_boss2_carry
    global boss3_carry
    global next_boss3_carry
    global boss4_carry
    global next_boss4_carry
    global boss5_carry
    global next_boss5_carry
    global next_boss1_total
    global next_boss2_total
    global next_boss3_total
    global next_boss4_total
    global next_boss5_total
    global mylist_end
    global mylist_start

    is_day_rolled = True
    while True:
        # 5時ロールオーバー処理
        if datetime.now(JST).strftime('%H:%M:%S') < '04:59:59':
            is_day_rolled = False
        if datetime.now(JST).strftime('%H:%M:%S') >= '04:59:59' and not is_day_rolled:
            boss1_mage = []
            boss1_physic = []
            boss1_carry = []
            boss2_mage = []
            boss2_physic = []
            boss2_carry = []
            boss3_mage = []
            boss3_physic = []
            boss3_carry = []
            boss4_mage = []
            boss4_physic = []
            boss4_carry = []
            boss4_carry = []
            boss5_mage = []
            boss5_physic = []
            boss5_carry = []
            boss1_total = []
            boss2_total = []
            boss3_total = []
            boss4_total = []
            boss5_total = []
            next_boss1_mage = []
            next_boss1_physic = []
            next_boss1_carry = []
            next_boss1_total = []
            next_boss2_mage = []
            next_boss2_physic = []
            next_boss2_carry = []
            next_boss2_total = []
            next_boss3_mage = []
            next_boss3_physic = []
            next_boss3_carry = []
            next_boss3_total = []
            next_boss4_mage = []
            next_boss4_physic = []
            next_boss4_carry = []
            next_boss4_total = []
            next_boss5_mage = []
            next_boss5_physic = []
            next_boss5_carry = []
            next_boss5_total = []
            guild = client.get_guild(lot_server_id)
            role = discord.utils.get(guild.roles,name='タスキル')
            role_toroi = discord.utils.get(guild.roles,name='トロイメライ')
            for member in guild.members:
                if role_toroi in member.roles:
                    if not member.display_name in mylist_start:
                        mylist_start.append(member.display_name)
                    if member.display_name in mylist_end:
                        mylist_end.remove(member.display_name)
                if role in member.roles:
                    await member.remove_roles(role)
            is_day_rolled = True
        await asyncio.sleep(1)

# 起動時に動作する処理
@client.event
async def on_ready():
    global ID_CHANNEL_RESERV
    global playData
    global ClabatoMode

    # 起動したらターミナルにログイン通知が表示される
    print('ログインしました')
    #if ID_CHANNEL_RESERV is not None:
        #await init_reserv_channel()
    #guild = client.get_guild(lot_server_id)
    #for member in guild.members:
        #playData.append(PlayerData(member, 0, False))
    asyncio.ensure_future(rollover_by5am())

mylist_start = []
mylist_end = []

@client.event
async def on_message(message):

    msg_content = jaconv.normalize(message.content)

    #ここで書き換えるものだけ。参照するだけのグローバル変数はここに書かなくて良い
    global bossindex
    global round
    global boss1_mage
    global boss1_physic
    global boss2_mage
    global boss2_physic
    global boss3_mage
    global boss3_physic
    global boss4_mage
    global boss4_physic
    global boss5_mage
    global boss5_physic
    global boss1_total
    global boss2_total
    global boss3_total
    global boss4_total
    global boss5_total
    global next_boss1_mage
    global next_boss1_physic
    global next_boss2_mage
    global next_boss2_physic
    global next_boss3_mage
    global next_boss3_physic
    global next_boss4_mage
    global next_boss4_physic
    global next_boss5_mage
    global next_boss5_physic
    global boss1_carry
    global next_boss1_carry
    global boss2_carry
    global next_boss2_carry
    global boss3_carry
    global next_boss3_carry
    global boss4_carry
    global next_boss4_carry
    global boss5_carry
    global next_boss5_carry
    global next_boss1_total
    global next_boss2_total
    global next_boss3_total
    global next_boss4_total
    global next_boss5_total
    global emoji_owari
    global playData
    global ClabatoMode
    global mayopote
    global dennys
    orig_user = message.author.display_name

    # メッセージ送信者がBotだった場合は無視する
    if message.author.bot:
        return
    #★予約状況の表示★
    if message.content == '凸予約状況':
        text = "```{round2}周目一覧\n\
 ==========【ワイバーン】==========\n\
 1ボス物理 \n{boss1_1} \n 1ボス魔法 \n{boss1_2}\n 1ボス持越し \n{boss1_3}\n\
 ==========【ランドスロース】==========\n\
 2ボス物理 \n{boss2_1} \n 2ボス魔法 \n{boss2_2}\n 2ボス持越し \n{boss2_3}\n\
 ==========【ムシュフシュ】==========\n\
 3ボス物理 \n{boss3_1} \n 3ボス魔法 \n{boss3_2}\n 3ボス持越し \n{boss3_3}\n\
 ==========【ティタノタートル】==========\n\
 4ボス物理 \n{boss4_1} \n 4ボス魔法 \n{boss4_2}\n 4ボス持越し \n{boss4_3}\n\
 ==========【オルレオン】==========\n\
 5ボス物理 \n{boss5_1} \n 5ボス魔法 \n{boss5_2}\n 5ボス持越し \n{boss5_3}\
            ```"
        result = text.format(round2=round,\
            boss1_1=boss1_physic,boss1_2=boss1_mage,boss1_3=boss1_carry,\
            boss2_1=boss2_physic,boss2_2=boss2_mage,boss2_3=boss2_carry,\
            boss3_1=boss3_physic,boss3_2=boss3_mage,boss3_3=boss3_carry,\
            boss4_1=boss4_physic,boss4_2=boss4_mage,boss4_3=boss4_carry,\
            boss5_1=boss5_physic,boss5_2=boss5_mage,boss5_3=boss5_carry)
        await message.channel.send(result)
    if message.content == '次凸予約状況':
        text = "```{round2}周目一覧\n\
 ==========【ワイバーン】==========\n\
 1ボス物理 \n{boss1_1} \n 1ボス魔法 \n{boss1_2}\n 1ボス持越し \n{boss1_3}\n\
 ==========【ランドスロース】==========\n\
 2ボス物理 \n{boss2_1} \n 2ボス魔法 \n{boss2_2}\n 2ボス持越し \n{boss2_3}\n\
 ==========【ムシュフシュ】==========\n\
 3ボス物理 \n{boss3_1} \n 3ボス魔法 \n{boss3_2}\n 3ボス持越し \n{boss3_3}\n\
 ==========【ティタノタートル】==========\n\
 4ボス物理 \n{boss4_1} \n 4ボス魔法 \n{boss4_2}\n 4ボス持越し \n{boss4_3}\n\
 ==========【オルレオン】==========\n\
 5ボス物理 \n{boss5_1} \n 5ボス魔法 \n{boss5_2}\n 5ボス持越し \n{boss5_3}\
            ```"
        result = text.format(round2=round+1,\
            boss1_1=next_boss1_physic,boss1_2=next_boss1_mage,boss1_3=next_boss1_carry,\
            boss2_1=next_boss2_physic,boss2_2=next_boss2_mage,boss2_3=next_boss2_carry,\
            boss3_1=next_boss3_physic,boss3_2=next_boss3_mage,boss3_3=next_boss3_carry,\
            boss4_1=next_boss4_physic,boss4_2=next_boss4_mage,boss4_3=next_boss4_carry,\
            boss5_1=next_boss5_physic,boss5_2=next_boss5_mage,boss5_3=next_boss5_carry)
        await message.channel.send(result)
    #予約取り消し
    if message.content == '予約取消' or message.content == '予約取消し' or message.content == '予約取り消し':
        for p in playData:
            if p.user.name == message.author.name:
                if p.user.name in boss1_physic:
                    boss1_physic.remove(p.user.name)
                if p.user.name in boss1_mage:
                    boss1_mage.remove(p.user.name)
                if p.user.name in boss1_carry:
                    boss1_carry.remove(p.user.name)
                if p.user.name in boss1_total:
                    boss1_total.remove(p.user.name)
                if p.user.name in boss2_physic:
                    boss2_physic.remove(p.user.name)
                if p.user.name in boss2_mage:
                    boss2_mage.remove(p.user.name)
                if p.user.name in boss2_carry:
                    boss2_carry.remove(p.user.name)
                if p.user.name in boss2_total:
                    boss2_total.remove(p.user.name)
                if p.user.name in boss3_physic:
                    boss3_physic.remove(p.user.name)
                if p.user.name in boss3_mage:
                    boss3_mage.remove(p.user.name)
                if p.user.name in boss3_carry:
                    boss3_carry.remove(p.user.name)
                if p.user.name in boss3_total:
                    boss3_total.remove(p.user.name)
                if p.user.name in boss4_physic:
                    boss4_physic.remove(p.user.name)
                if p.user.name in boss4_mage:
                    boss4_mage.remove(p.user.name)
                if p.user.name in boss4_carry:
                    boss4_carry.remove(p.user.name)
                if p.user.name in boss4_total:
                    boss4_total.remove(p.user.name)
                if p.user.name in boss5_physic:
                    boss5_physic.remove(p.user.name)
                if p.user.name in boss5_mage:
                    boss5_mage.remove(p.user.name)
                if p.user.name in boss5_carry:
                    boss5_carry.remove(p.user.name)
                if p.user.name in boss5_total:
                    boss5_total.remove(p.user.name)
                if p.user.name in next_boss1_physic:
                    next_boss1_physic.remove(p.user.name)
                if p.user.name in next_boss1_mage:
                    next_boss1_mage.remove(p.user.name)
                if p.user.name in next_boss1_carry:
                    next_boss1_carry.remove(p.user.name)
                if p.user.name in next_boss1_total:
                    next_boss1_total.remove(p.user.name)
                if p.user.name in next_boss2_physic:
                    next_boss2_physic.remove(p.user.name)
                if p.user.name in next_boss2_mage:
                    next_boss2_mage.remove(p.user.name)
                if p.user.name in next_boss2_carry:
                    next_boss2_carry.remove(p.user.name)
                if p.user.name in next_boss2_total:
                    next_boss2_total.remove(p.user.name)
                if p.user.name in next_boss3_physic:
                    next_boss3_physic.remove(p.user.name)
                if p.user.name in next_boss3_mage:
                    next_boss3_mage.remove(p.user.name)
                if p.user.name in next_boss3_carry:
                    next_boss3_carry.remove(p.user.name)
                if p.user.name in next_boss3_total:
                    next_boss3_total.remove(p.user.name)
                if p.user.name in next_boss4_physic:
                    next_boss4_physic.remove(p.user.name)
                if p.user.name in next_boss4_mage:
                    next_boss4_mage.remove(p.user.name)
                if p.user.name in next_boss4_carry:
                    next_boss4_carry.remove(p.user.name)
                if p.user.name in next_boss4_total:
                    next_boss4_total.remove(p.user.name)
                if p.user.name in next_boss5_physic:
                    next_boss5_physic.remove(p.user.name)
                if p.user.name in next_boss5_mage:
                    next_boss5_mage.remove(p.user.name)
                if p.user.name in next_boss5_carry:
                    next_boss5_carry.remove(p.user.name)
                if p.user.name in next_boss5_total:
                    next_boss5_total.remove(p.user.name)
                reply = f'{message.author.mention} 自分の予約を全て取り消したよ。もう一回入れてね'
                await message.channel.send(reply)

    #予約リストの削除
    if message.content.startswith('/代マイ一物'):
        guild = client.get_guild(lot_server_id)
        role = discord.utils.get(guild.roles,name='トロイメライ')       
        for member in guild.members:
            if role in member.roles:
                getMoji = message.mentions[0].id
                getMoji2 = guild.get_member(getMoji)
                if getMoji2 in guild.members:
                    if getMoji2.display_name in boss1_physic:
                        boss1_physic.remove(getMoji2.display_name)
                        reply = f'{message.author.mention} {getMoji2.display_name}を予約一覧から削除したよ'
                        await message.channel.send(reply)
                        break
                    else:
                        reply = f'{message.author.mention} {getMoji2.display_name}は完了一覧に存在しないよ'
                        await message.channel.send(reply)
                        break
    #周回数更新
    if re.match(r'^周回数', msg_content):
        round = msg_content[4:]
        round = int(round)
        reply = f'{message.author.mention} 周回数を更新しました。{round}周目となりました'
        await message.channel.send(reply)
    #予約リストの消去＋次ボス周回のカウントアップ
    if message.content == 'オールクリア':
        round = round + 1
        boss1_mage = []
        boss1_physic = []
        boss1_carry = []
        boss2_mage = []
        boss2_physic = []
        boss2_carry = []
        boss3_mage = []
        boss3_physic = []
        boss3_carry = []
        boss4_mage = []
        boss4_physic = []
        boss4_carry = []
        boss4_carry = []
        boss5_mage = []
        boss5_physic = []
        boss5_carry = []
        boss1_total = []
        boss2_total = []
        boss3_total = []
        boss4_total = []
        boss5_total = []
        boss1_mage.extend(next_boss1_mage)
        boss1_physic.extend(next_boss1_physic)
        boss1_carry.extend(next_boss1_carry)
        boss1_total.extend(next_boss1_total)
        next_boss1_mage = []
        next_boss1_physic = []
        next_boss1_carry = []
        next_boss1_total = []
        boss2_mage.extend(next_boss2_mage)
        boss2_physic.extend(next_boss2_physic)
        boss2_carry.extend(next_boss2_carry)
        boss2_total.extend(next_boss2_total)
        next_boss2_mage = []
        next_boss2_physic = []
        next_boss2_carry = []
        next_boss2_total = []
        boss3_mage.extend(next_boss3_mage)
        boss3_physic.extend(next_boss3_physic)
        boss3_carry.extend(next_boss3_carry)
        boss3_total.extend(next_boss3_total)
        next_boss3_mage = []
        next_boss3_physic = []
        next_boss3_carry = []
        next_boss3_total = []
        boss4_mage.extend(next_boss4_mage)
        boss4_physic.extend(next_boss4_physic)
        boss4_carry.extend(next_boss4_carry)
        boss4_total.extend(next_boss4_total)
        next_boss4_mage = []
        next_boss4_physic = []
        next_boss4_carry = []
        next_boss4_total = []
        boss5_mage.extend(next_boss5_mage)
        boss5_physic.extend(next_boss5_physic)
        boss5_carry.extend(next_boss5_carry)
        boss5_total.extend(next_boss5_total)
        next_boss5_mage = []
        next_boss5_physic = []
        next_boss5_carry = []
        next_boss5_total = []
        reply = f'{message.author.mention} 現在の予約リストをリセットし、次週分を持ち越しました'
        await message.channel.send(reply)

    # 3凸完了した時にリストに登録する
    if message.channel.id == lot_channel_id:
        guild = client.get_guild(lot_server_id)
        role = discord.utils.get(guild.roles,name='トロイメライ')       
        if role in message.author.roles:
            sleep(1)
            if not message.author.display_name in mylist_end:
                mylist_start.remove(message.author.display_name)
                mylist_end.append(message.author.display_name)
    #完了・未完了一覧の呼び出し
    if message.content == 'メンバー':
        guild = client.get_guild(lot_server_id)
        text = "```3凸未完者 \n{name} \n\n3凸完了者 \n{name2}```"
        result = text.format(name=mylist_start,name2=mylist_end)
        await message.channel.send(result)
    #タスキル設定
    if message.content == 'タスキル':
        guild = client.get_guild(lot_server_id)
        role = discord.utils.get(message.guild.roles, name='タスキル')
        if role in message.author.roles:
            await message.author.remove_roles(role)
            reply = f'{message.author.mention} タスキル解除したよ'
            await message.channel.send(reply)
        else:
            await message.author.add_roles(role)
            reply = f'{message.author.mention} タスキル設定したよ'
            await message.channel.send(reply)
    # 「/ping」と発言したら返事が返る処理
    if message.content == '/ping':
        await message.channel.send('○ ˶˙・˙˶ ○pongdering!')
    if message.content == 'まよぽてちは？':
        await message.channel.send('https://twitter.com/SINoPotato/status/1042335422628519937?s=20')
    if message.content == 'まよぽてち変身':
        await message.channel.send('https://twitter.com/SINoPotato/status/1235580352162197504?s=20')
    if message.content == 'smt':
        text = "₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　 ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾\n\
\n\
₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾    𝐒𝐔𝐏𝐄𝐑 𝐌𝐀𝐘𝐎𝐏𝐎𝐓𝐄 𝐓𝐈𝐌𝐄    ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾\n\
\n\
₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾ 　 ₍₍⁽⁽" + str(client.get_emoji(mayopote)) + " ₎₎⁾⁾\n\
            "
        await message.channel.send(text)   
    if message.content == 'sdt':
        text = "₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　 ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾\n\
\n\
₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾    𝐒𝐔𝐏𝐄𝐑 𝐃𝐄𝐍𝐍𝐘'𝐒 𝐓𝐈𝐌𝐄          ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾\n\
\n\
₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　  ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾ 　 ₍₍⁽⁽" + str(client.get_emoji(dennys)) + " ₎₎⁾⁾\n\
            "
        await message.channel.send(text) 
    #コマンドにて完了一覧から削除し、未完了一覧に登録する
    if message.content.startswith('/代理削除'):
        guild = client.get_guild(lot_server_id)
        role = discord.utils.get(guild.roles,name='トロイメライ')       
        for member in guild.members:
            if role in member.roles:
                getMoji = message.mentions[0].id
                getMoji2 = guild.get_member(getMoji)
                if getMoji2 in guild.members:
                    if getMoji2.display_name in mylist_end:
                        mylist_start.append(getMoji2.display_name)
                        mylist_end.remove(getMoji2.display_name)
                        reply = f'{message.author.mention} {getMoji2.display_name}を完了一覧から削除したよ'
                        await message.channel.send(reply)
                        break
                    else:
                        reply = f'{message.author.mention} {getMoji2.display_name}は完了一覧に存在しないよ'
                        await message.channel.send(reply)
                        break
    #コマンドにて未完了一覧から削除し、完了一覧に登録する
    if message.content.startswith('/代理登録'):
        guild = client.get_guild(lot_server_id)
        role = discord.utils.get(guild.roles,name='トロイメライ')       
        for member in guild.members:
            if role in member.roles:
                getMoji = message.mentions[0].id
                getMoji2 = guild.get_member(getMoji)
                if getMoji2 in guild.members:
                    if  getMoji2.display_name in mylist_start:
                        mylist_start.remove(getMoji2.display_name)
                        mylist_end.append(getMoji2.display_name)
                        reply = f'{message.author.mention} {getMoji2.display_name}を完了一覧に登録したよ'
                        await message.channel.send(reply)
                        break
                    else:
                        reply = f'{message.author.mention} {getMoji2.display_name}は既に終わっているよ'
                        await message.channel.send(reply)
                        break
                else:
                    print('いないよ')
    #コマンドにて未完了一覧から削除し、完了一覧に登録する
    if message.content.startswith('/リセット'):
        guild = client.get_guild(lot_server_id)
        role_toroi = discord.utils.get(guild.roles,name='トロイメライ')
        for member in guild.members:
            if role_toroi in member.roles:
                if not member.display_name in mylist_start:
                    mylist_start.append(member.display_name)
                if member.display_name in mylist_end:
                    mylist_end.remove(member.display_name)
    # 持越時間予想
    if re.match(r'^持越時間|^持越し時間|^持ち越し時間|^rollover|^ro', msg_content):
        await rollover_simulate(message, msg_content)
        return
    #クラバトモード起動
    if message.content == 'システムコール クラバトモード':
        ClabatoMode = True
        playData = []
        guild = client.get_guild(lot_server_id)
        for member in guild.members:
            playData.append(PlayerData(member, 0, False))
        #if ID_CHANNEL_RESERV is not None:
            #await init_reserv_channel()
        reply = f'{message.author.mention} クラバトモードを開始します'
        await message.channel.send(reply)
    elif message.content == 'システムコール ノーマルモード':
        ClabatoMode = False
        reply = f'{message.author.mention} クラバトモードを停止します'
        await message.channel.send(reply)
    
#    if re.match(r'^検索', msg_content):
#        get1totu = []
#        get2totukari = []
#        get2totu = []
#        get3totukari = []
#        get3totu = []
#        DaburiCount = 0
#        getNo = msg_content[3:]
#        cell = worksheet.find(getNo)
#        cell.row
#        getGyo1 = int(cell.row)
#        for num1 in range(15,20):
#            appendchara = worksheet.cell(getGyo1, num1).value
#            get1totu.append(appendchara)
#        appendNo0 =worksheet.cell(getGyo1, 1).value
#        print(appendNo0,get1totu)
#        for num2 in range(5,38):
#            for num1 in range(15,20):
#                appendchara =worksheet.cell(num2, num1).value
#                get2totukari.append(appendchara)
#            if get2totukari[0] == "":
#                get2totukari = []
#                continue
#            for num3 in range(0,5):
#                num4 = 0
#                while num4 < 5:
#                    if get2totukari[num3] == get1totu[num4]:
#                        DaburiCount += 1
#                        break
#                    else:
#                        num4 += 1
#            if DaburiCount <= 2:
#                get2totu.extend(get2totukari)
#                appendNo1 =worksheet.cell(num2, 1).value
#                DaburiCount = 0
#                break
#            else:
#                DaburiCount = 0
#                get2totukari = []
#                continue
#        for num6 in range(5,38):
#            for num5 in range(15,20):
#                appendchara =worksheet.cell(num6, num5).value
#                get3totukari.append(appendchara)
#            if get3totukari[0] == "":
#                get3totukari = []
#                continue
#            for num7 in range(0,5):
#                num8 = 0
#                num9 = 0
#                while num8 < 5:
#                    if get3totukari[num7] == get1totu[num8]:
#                        DaburiCount += 1
#                        break
#                    else:
#                        num8 += 1
#                while num9 < 5:
#                    if get3totukari[num7] == get2totu[num9]:
#                        DaburiCount += 1
#                        break
#                    else:
#                        num9 += 1
#            if DaburiCount <= 1:
#                get3totu.extend(get3totukari)
#                appendNo2 =worksheet.cell(num6, 1).value
#                print(appendNo1,get2totu)
#                print(appendNo2,get3totu)
#                DaburiCount = 0
#                break
#            else:
#                DaburiCount = 0
#                get3totukari = []
#                continue


    #★★★以下廃止機能★★★
    # ★ダメージ情報の取得★
#    if message.channel.id == damage_channel_id:
#        for p in playData:
#            if p.user == orig_user:
#                if bossindex == 0:
#                    if p.user in boss1_total:
#                        if len(message.content) <= 4:
#                            str_boss_damage = message.content
#                            int_boss_damage = int(str_boss_damage)
#                            p.add_Boss_Damage(int_boss_damage)
#                        else:
#                            await message.channel.send(f"{message.author.mention} 4桁以下ダメージで入れてね(1000以下は切り捨てだからね)")
#                    else:
#                        await message.channel.send(f"{message.author.mention} ややこしいから記入しちゃだめよ")
#                elif bossindex == 1:
#                    if p.user.name in boss2_total: 
#                        if len(message.author.mention) <= 4:
#                            str_boss_damage = message.author.mention
#                            int_boss_damage = int(str_boss_damage)
#                            p.add_Boss_Damage(int_boss_damage)
#                        else:
#                            await message.channel.send(f"{message.author.mention} 4桁以下ダメージで入れてね(1000以下は切り捨てだからね)")
#                    else:
#                        await message.channel.send(f"{message.author.mention} ややこしいから記入しちゃだめよ")
#                elif bossindex == 2:
#                    if p.user.name in boss3_total: 
#                        if len(orig_user.mention) <= 4:
#                            str_boss_damage = orig_user.mention
#                            int_boss_damage = int(str_boss_damage)
#                            p.add_Boss_Damage(int_boss_damage)
#                        else:
#                            await message.channel.send(f"{message.author.mention} 4桁以下ダメージで入れてね(1000以下は切り捨てだからね)")
#                    else:
#                        await message.channel.send(f"{message.author.mention} ややこしいから記入しちゃだめよ")
#                elif bossindex == 3:
#                    if p.user.name in boss4_total: 
#                        if len(orig_user.mention) <= 4:
#                            str_boss_damage = orig_user.mention
#                            int_boss_damage = int(str_boss_damage)
#                            p.add_Boss_Damage(int_boss_damage)
#                        else:
#                            await message.channel.send(f"{message.author.mention} 4桁以下ダメージで入れてね(1000以下は切り捨てだからね)")
#                    else:
#                        await message.channel.send(f"{message.author.mention} ややこしいから記入しちゃだめよ")
#                elif bossindex == 4:
#                    if p.user.name in boss5_total: 
#                        if len(orig_user.mention) <= 4:
#                            str_boss_damage = orig_user.mention
#                            int_boss_damage = int(str_boss_damage)
#                            p.add_Boss_Damage(int_boss_damage)
#                        else:
#                            await message.channel.send(f"{message.author.mention} 4桁以下ダメージで入れてね(1000以下は切り捨てだからね)")
#                    else:
#                        await message.channel.send(f"{message.author.mention} ややこしいから記入しちゃだめよ")

    # ★メンバー情報、ボス情報をすべて初期化する。メンバーに変更があった場合に行って下さい★
#    if message.channel.id == input_channel_id:
#        if message.content == '/同時凸結果':
#            hp = bosshp[bossindex]
#            for p in playData:
#                if p.Boss_Confirm_Flag == True:
#                    hp = hp - p.Boss_Damage
#            for p in playData:
#                if p.Boss_Confirm_Flag == False:
#                    if p.Boss_Damage != 0:
#                        rolled_time = 90 * (1 - hp / p.Boss_Damage) + 20
#                        if rolled_time >= 90:
#                            rolled_time = 90
#                        reply = f"{p.user}さんの予想される持越時間は" + str(rolled_time) + "秒です"
#                        await message.channel.send(reply)
    #タスキル設定
    #if message.content == '終わり':
        #guild = client.get_guild(lot_server_id)
        #role = discord.utils.get(message.guild.roles, name='タスキル')
        #await message.author.remove_roles(role)
        #reply = f'{message.author.mention} タスキル解除したよ'
        #await message.channel.send(reply)


# 30秒に一回ループし、5時になったらタスキルロールの削除(廃止)
@tasks.loop(seconds=30)
async def time_check():
    global OutputCount
    global yoyaku_id
    global round
    global boss1_mage
    global boss1_physic
    global boss2_mage
    global boss2_physic
    global boss3_mage
    global boss3_physic
    global boss4_mage
    global boss4_physic
    global boss5_mage
    global boss5_physic
    global boss1_total
    global boss2_total
    global boss3_total
    global boss4_total
    global boss5_total
    global next_boss1_mage
    global next_boss1_physic
    global next_boss2_mage
    global next_boss2_physic
    global next_boss3_mage
    global next_boss3_physic
    global next_boss4_mage
    global next_boss4_physic
    global next_boss5_mage
    global next_boss5_physic
    global boss1_carry
    global next_boss1_carry
    global boss2_carry
    global next_boss2_carry
    global boss3_carry
    global next_boss3_carry
    global boss4_carry
    global next_boss4_carry
    global boss5_carry
    global next_boss5_carry
    global next_boss1_total
    global next_boss2_total
    global next_boss3_total
    global next_boss4_total
    global next_boss5_total

    if OutputCount >= 60:
        yoyaku_channel = client.get_channel(yoyaku_id)
        text = "```{round2}周目一覧\n\
 ==========【ワイバーン】==========\n\
 1ボス物理 \n{boss1_1} \n 1ボス魔法 \n{boss1_2}\n 1ボス持越し \n{boss1_3}\n\
 ==========【ランドスロース】==========\n\
 2ボス物理 \n{boss2_1} \n 2ボス魔法 \n{boss2_2}\n 2ボス持越し \n{boss2_3}\n\
 ==========【ムシュフシュ】==========\n\
 3ボス物理 \n{boss3_1} \n 3ボス魔法 \n{boss3_2}\n 3ボス持越し \n{boss3_3}\n\
 ==========【ティタノタートル】==========\n\
 4ボス物理 \n{boss4_1} \n 4ボス魔法 \n{boss4_2}\n 4ボス持越し \n{boss4_3}\n\
 ==========【オルレオン】==========\n\
 5ボス物理 \n{boss5_1} \n 5ボス魔法 \n{boss5_2}\n 5ボス持越し \n{boss5_3}\
            ```"
        result = text.format(round2=round,\
            boss1_1=boss1_physic,boss1_2=boss1_mage,boss1_3=boss1_carry,\
            boss2_1=boss2_physic,boss2_2=boss2_mage,boss2_3=boss2_carry,\
            boss3_1=boss3_physic,boss3_2=boss3_mage,boss3_3=boss3_carry,\
            boss4_1=boss4_physic,boss4_2=boss4_mage,boss4_3=boss4_carry,\
            boss5_1=boss5_physic,boss5_2=boss5_mage,boss5_3=boss5_carry)
        await yoyaku_channel.send(result)
        OutputCount = 0
    else:
        OutputCount = OutputCount + 1

#ループ処理実行
#time_check.start()

# Botの起動とDiscordサーバーへの接続
client.run(TOKEN)
